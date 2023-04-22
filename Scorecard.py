
from datetime import datetime
start_time = datetime.now()
import openpyxl
import pandas as pd
import os
from datetime import datetime
start_time = datetime.now()
 #opening pakistan inning file
Pakistan_Inning = open("pak_inns1.txt","r+") 
 #opening india inning file
India_Inning = open("india_inns2.txt","r+")
#reading teams text file
teams = open("teams.txt","r+")
team = teams.readlines()
 
Pakistan_team = team[0]
#spliting at ','
pakistan_players = Pakistan_team[23:-1:].split(",")
#spliting at ','
India_team = team[2]
India_players = India_team[20:-1:].split(",")


India  = India_Inning.readlines() 
for i in India:         #removing line space
    if i=='\n':
        India.remove(i)
      

Pakistan = Pakistan_Inning.readlines() 
for i in Pakistan:
    if i=='\n':                   #removing line space
        Pakistan.remove(i)

wb = openpyxl.Workbook()
ws = wb.active                


India_Fall_of_wickets=0
Pakistan_fall_of_wickets=0
Pakistan_Byes=0                 #created new variables for counting
Pakistan_bowlers_total=0


Out_Pakistan_batsman={}
India_bowlers={}
India_bats={}              #created empty dictionary
Pakistan_bats={}
Pakistan_bowlers={}



try:
    for l in Pakistan:
        x=l.index(".")
        Pak_over =l[0:x+2]
        temp=l[x+2::].split(",")
        Current_Ball=temp[0].split("to") #0 2

        if f"{Current_Ball[0].strip()}" not in India_bowlers.keys() :
            India_bowlers[f"{Current_Ball[0].strip()}"]=[1,0,0,0,0,0,0]   
        elif "wide" in temp[1]:
            pass
        elif "bye" in temp[1]:
            if "FOUR" in temp[2]:
                Pakistan_Byes+=4
            elif "1" in temp[2]:
                Pakistan_Byes+=1            #over0,
            elif "2" in temp[2]:            #NB4
                Pakistan_Byes+=2            #medan1
            elif "3" in temp[2]:            #runs2
                Pakistan_Byes+=3            #Wickets3,
            elif "4" in temp[2]:            #ECO6
                Pakistan_Byes+=4            # WD5, 
            elif "5" in temp[2]:
                Pakistan_Byes+=5

        else:
            India_bowlers[f"{Current_Ball[0].strip()}"][0]+=1
        
        if f"{Current_Ball[1].strip()}" not in Pakistan_bats.keys() and temp[1]!="wide":
            Pakistan_bats[f"{Current_Ball[1].strip()}"]=[0,1,0,0,0]                #runs ,ball ,4s ,6s , sr
        elif "wide" in temp[1] :
            pass
        else:
            Pakistan_bats[f"{Current_Ball[1].strip()}"][1]+=1
        

        if "out" in temp[1]:
            India_bowlers[f"{Current_Ball[0].strip()}"][3]+=1
            if "Bowled" in temp[1].split("!!")[0]:
                Out_Pakistan_batsman[f"{Current_Ball[1].strip()}"]=("b" + Current_Ball[0])
            elif "Caught" in temp[1].split("!!")[0]:
                w=(temp[1].split("!!")[0]).split("by")                                                        #this is for wicket criteria 
                Out_Pakistan_batsman[f"{Current_Ball[1].strip()}"]=("c" + w[1] +" b " + Current_Ball[0])
            elif "Lbw" in temp[1].split("!!")[0]:
                Out_Pakistan_batsman[f"{Current_Ball[1].strip()}"]=("lbw  b "+Current_Ball[0])

        

        if "no run" in temp[1] or "out" in temp[1] :
            India_bowlers[f"{Current_Ball[0].strip()}"][2]+=0
            Pakistan_bats[f"{Current_Ball[1].strip()}"][0]+=0
        elif "1 run" in temp[1]:
            India_bowlers[f"{Current_Ball[0].strip()}"][2]+=1
            Pakistan_bats[f"{Current_Ball[1].strip()}"][0]+=1
        elif "2 run" in temp[1]:
            India_bowlers[f"{Current_Ball[0].strip()}"][2]+=2
            Pakistan_bats[f"{Current_Ball[1].strip()}"][0]+=2
        elif "3 run" in temp[1]:
            India_bowlers[f"{Current_Ball[0].strip()}"][2]+=3          #this is for pakistan battin)and inia bowling counting
            Pakistan_bats[f"{Current_Ball[1].strip()}"][0]+=3
        elif "4 run" in temp[1]:
            India_bowlers[f"{Current_Ball[0].strip()}"][2]+=4
            Pakistan_bats[f"{Current_Ball[1].strip()}"][0]+=4
        elif "FOUR" in temp[1]:
            India_bowlers[f"{Current_Ball[0].strip()}"][2]+=4
            Pakistan_bats[f"{Current_Ball[1].strip()}"][0]+=4
            Pakistan_bats[f"{Current_Ball[1].strip()}"][2]+=1
        elif "SIX" in temp[1]:
            India_bowlers[f"{Current_Ball[0].strip()}"][2]+=6
            Pakistan_bats[f"{Current_Ball[1].strip()}"][0]+=6
            Pakistan_bats[f"{Current_Ball[1].strip()}"][3]+=1
        elif "wide" in temp[1]:
            if "wides" in temp[1]:
                India_bowlers[f"{Current_Ball[0].strip()}"][2]+=int(temp[1][1])
                India_bowlers[f"{Current_Ball[0].strip()}"][5]+=int(temp[1][1])
            else:
                India_bowlers[f"{Current_Ball[0].strip()}"][2]+=1
                India_bowlers[f"{Current_Ball[0].strip()}"][5]+=1
except:
    print('error1')
    exit()

for var in Pakistan_bats.values():
    var[-1]=round((var[0]/var[1])*100 , 2)




####this is same for india inning as for pakistan batting
try:
    India_bowlers_total=0
    ind_byes=0
    out_ind_bat={}
    for l in India:
        x=l.index(".")
        over_ind=l[0:x+2]

        temp=l[x+2::].split(",")

        Current_Ball=temp[0].split("to") #0 2
        if f"{Current_Ball[0].strip()}" not in Pakistan_bowlers.keys() :
            Pakistan_bowlers[f"{Current_Ball[0].strip()}"]=[1,0,0,0,0,0,0]   #over0, medan1, runs2, Wickets3, NB4, WD5, ECO6
        elif "wide" in temp[1]:
            pass
        elif "bye" in temp[1]:
            if "FOUR" in temp[2]:
                ind_byes+=4
            elif "1" in temp[2]:
                ind_byes+=1
            elif "2" in temp[2]:
                ind_byes+=2
            elif "3" in temp[2]:
                ind_byes+=3
            elif "4" in temp[2]:
                ind_byes+=4
            elif "5" in temp[2]:
                ind_byes+=5
        else:
            Pakistan_bowlers[f"{Current_Ball[0].strip()}"][0]+=1
        
        if f"{Current_Ball[1].strip()}" not in India_bats.keys() and temp[1]!="wide":
            India_bats[f"{Current_Ball[1].strip()}"]=[0,1,0,0,0] #[runs,ball,4s,6s,sr]
        elif "wide" in temp[1] :
            pass
        else:
            India_bats[f"{Current_Ball[1].strip()}"][1]+=1
        

        if "out" in temp[1]:
            Pakistan_bowlers[f"{Current_Ball[0].strip()}"][3]+=1
            
            if "Bowled" in temp[1].split("!!")[0]:
                out_ind_bat[f"{Current_Ball[1].strip()}"]=("b" + Current_Ball[0])
            elif "Caught" in temp[1].split("!!")[0]:
                w=(temp[1].split("!!")[0]).split("by")
                out_ind_bat[f"{Current_Ball[1].strip()}"]=("c" + w[1] +" b " + Current_Ball[0])
            elif "Lbw" in temp[1].split("!!")[0]:
                out_ind_bat[f"{Current_Ball[1].strip()}"]=("lbw  b "+Current_Ball[0])

        
        
        if "no run" in temp[1] or "out" in temp[1] :
            Pakistan_bowlers[f"{Current_Ball[0].strip()}"][2]+=0
            India_bats[f"{Current_Ball[1].strip()}"][0]+=0
        elif "1 run" in temp[1]:
            Pakistan_bowlers[f"{Current_Ball[0].strip()}"][2]+=1
            India_bats[f"{Current_Ball[1].strip()}"][0]+=1
        elif "2 run" in temp[1]:
            Pakistan_bowlers[f"{Current_Ball[0].strip()}"][2]+=2
            India_bats[f"{Current_Ball[1].strip()}"][0]+=2
        elif "3 run" in temp[1]:
            Pakistan_bowlers[f"{Current_Ball[0].strip()}"][2]+=3
            India_bats[f"{Current_Ball[1].strip()}"][0]+=3
        elif "4 run" in temp[1]:
            Pakistan_bowlers[f"{Current_Ball[0].strip()}"][2]+=4
            India_bats[f"{Current_Ball[1].strip()}"][0]+=4
        elif "FOUR" in temp[1]:
            Pakistan_bowlers[f"{Current_Ball[0].strip()}"][2]+=4
            India_bats[f"{Current_Ball[1].strip()}"][0]+=4
            India_bats[f"{Current_Ball[1].strip()}"][2]+=1
        elif "SIX" in temp[1]:
            Pakistan_bowlers[f"{Current_Ball[0].strip()}"][2]+=6
            India_bats[f"{Current_Ball[1].strip()}"][0]+=6
            India_bats[f"{Current_Ball[1].strip()}"][3]+=1
        elif "wide" in temp[1]:
            if "wides" in temp[1]:
                Pakistan_bowlers[f"{Current_Ball[0].strip()}"][2]+=int(temp[1][1])
                Pakistan_bowlers[f"{Current_Ball[0].strip()}"][5]+=int(temp[1][1])
            else:
                Pakistan_bowlers[f"{Current_Ball[0].strip()}"][2]+=1
                Pakistan_bowlers[f"{Current_Ball[0].strip()}"][5]+=1


    for var in India_bats.values():
        var[-1]=round((var[0]/var[1])*100 , 2)

    for var in Pakistan_bats.values():
        var[-1]=round((var[0]/var[1])*100 , 2)

    for var in India_bowlers.values():
        if var[0]%6==0:
            var[0] = var[0]//6
        else:
            var[0] = (var[0]//6) + (var[0]%6)/10

    for var in Pakistan_bowlers.values():    
        if var[0]%6==0:
            var[0] = var[0]//6
        else:
            var[0] = (var[0]//6) + (var[0]%6)/10

    for var in India_bowlers.values():    #economy of indian bowlers 
        x=str(var[0])
        if "." in x:
            balls = int(x[0])*6 + int(x[2])
            var[-1]=round((var[2]/balls)*6,1)
        else:
            var[-1] = round((var[2]/var[0]) ,1) 


    for var in Pakistan_bowlers.values(): #economy of indian bowlers
        x=str(var[0])
        if "." in x:
            balls = int(x[0])*6 + int(x[2])
            var[-1]=round((var[2]/balls)*6,1)
        else:
            var[-1] = round((var[2]/var[0]) ,1)


    Pak_Batters=[]
    for key in Pakistan_bats.keys():
        Pak_Batters.append(key)


    for i in range(len(Pakistan_bats)):
        ws.cell(5+i,1).value = Pak_Batters[i]
        ws.cell(5+i,5).value = Pakistan_bats[Pak_Batters[i]][0]
        ws.cell(5+i,6).value = Pakistan_bats[Pak_Batters[i]][1]
        ws.cell(5+i,7).value = Pakistan_bats[Pak_Batters[i]][2]
        ws.cell(5+i,8).value = Pakistan_bats[Pak_Batters[i]][3]   #printing the value for pakistan batting in cell
        ws.cell(5+i,9).value = Pakistan_bats[Pak_Batters[i]][4]
        if Pak_Batters[i] not in Out_Pakistan_batsman:
            ws.cell(5+i,3).value = "not out"
        else:
            ws.cell(5+i,3).value=Out_Pakistan_batsman[Pak_Batters[i]]
except:
    print('error2')
    exit()

ws.cell(3,1).value = "BATTERS"
ws["E3"] = "RUNS"
ws["F3"] = "BALLS"
ws["G3"] = " 4s "
ws["H3"] = " 6s "
ws["I3"] = "  SR  "


ws["A18"] = "BOWLER"
ws["C18"] = "OVER"
ws["D18"] = "MAIDEN"
ws["E18"] = "RUNS"
ws["F18"] = "WICKET"
ws["G18"] = "NO-BALL"
ws["H18"] = "WIDE"
ws["I18"] = "ECONOMY"

Pakistan_Bowlers=[]
for key in Pakistan_bowlers.keys():
    Pakistan_Bowlers.append(key)

for i in range(len(Pakistan_bowlers)):
    ws.cell(42+i,1).value = Pakistan_Bowlers[i]
    ws.cell(42+i,3).value = Pakistan_bowlers[Pakistan_Bowlers[i]][0]
    ws.cell(42+i,4).value = Pakistan_bowlers[Pakistan_Bowlers[i]][1]
    ws.cell(42+i,5).value = Pakistan_bowlers[Pakistan_Bowlers[i]][2]    #printing value for pakistan bowling 
    ws.cell(42+i,6).value = Pakistan_bowlers[Pakistan_Bowlers[i]][3]
    ws.cell(42+i,7).value = Pakistan_bowlers[Pakistan_Bowlers[i]][4]
    ws.cell(42+i,8).value = Pakistan_bowlers[Pakistan_Bowlers[i]][5]
    ws.cell(42+i,9).value = Pakistan_bowlers[Pakistan_Bowlers[i]][6]
    Pakistan_bowlers_total+=Pakistan_bowlers[Pakistan_Bowlers[i]][2]
    India_Fall_of_wickets+=Pakistan_bowlers[Pakistan_Bowlers[i]][3]


ws.cell(11+len(Pakistan_bats)+len(Pakistan_bowlers),1).value = "# INDIA"
ws.cell(11+len(Pakistan_bats)+len(Pakistan_bowlers),2).value = " INNINGS"

Indian_Batters=[]
for key in India_bats.keys():
    Indian_Batters.append(key)


for i in range(len(India_bats)):
    ws.cell(31+i,1).value = Indian_Batters[i]
    ws.cell(31+i,5).value = India_bats[Indian_Batters[i]][0]
    ws.cell(31+i,6).value = India_bats[Indian_Batters[i]][1]  #printing value of indian batters
    ws.cell(31+i,7).value = India_bats[Indian_Batters[i]][2]
    ws.cell(31+i,8).value = India_bats[Indian_Batters[i]][3]
    ws.cell(31+i,9).value = India_bats[Indian_Batters[i]][4]

    if Indian_Batters[i] not in out_ind_bat:
        ws.cell(31+i,3).value = "not out"
    else:
        ws.cell(31+i,3).value=out_ind_bat[Indian_Batters[i]]

ws["A29"] = "BATTERS"
ws["E29"] = "RUNS"
ws["F29"] = "BALLS"
ws["G29"] = " 4s "
ws["H29"] = " 6s "
ws["I29"] = "  SR  "


ws["A40"] = "BOWLER"
ws["C40"] = "OVER"
ws["D40"] = "MAIDEN"
ws["E40"] = "RUNS"
ws["F40"] = "WICKET"
ws["G40"] = "NO-BALL"
ws["H40"] = "WIDE"
ws["I40"] = "ECONOMY"

Indian_Bowlers=[]
for key in India_bowlers.keys():
    Indian_Bowlers.append(key)

for i in range(len(India_bowlers)):

    ws.cell(20+i,1).value = Indian_Bowlers[i]
    ws.cell(20+i,3).value = India_bowlers[Indian_Bowlers[i]][0]
    ws.cell(20+i,4).value = India_bowlers[Indian_Bowlers[i]][1]
    ws.cell(20+i,5).value = India_bowlers[Indian_Bowlers[i]][2]  
    ws.cell(20+i,6).value = India_bowlers[Indian_Bowlers[i]][3]     #printing value of indian bollers
    ws.cell(20+i,7).value = India_bowlers[Indian_Bowlers[i]][4]
    ws.cell(20+i,8).value = India_bowlers[Indian_Bowlers[i]][5]
    ws.cell(20+i,9).value = India_bowlers[Indian_Bowlers[i]][6]
    India_bowlers_total+=India_bowlers[Indian_Bowlers[i]][2]
    Pakistan_fall_of_wickets+=India_bowlers[Indian_Bowlers[i]][3]

India_total_Score=India_bowlers_total+Pakistan_Byes
Pakistan_Total_Score = Pakistan_bowlers_total+ind_byes




ws["E27"] = " "+str(India_total_Score) +" - " + str(India_Fall_of_wickets)
ws["F27"] = str(over_ind)
Eone=" "+str(Pakistan_Total_Score) +" - " + str(Pakistan_fall_of_wickets)
Fone = str(Pak_over)

wb.save("Scorecard.xlsx")

df = pd.read_excel('Scorecard.xlsx')

df = df.set_axis(['PAKISTAN', ' INNINGS'] + [" "," ",Eone,Fone," "," "," "], axis='columns')

df.to_excel('Scorecard.xlsx',index=False)

#This shall be the last lines of the code.
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))
